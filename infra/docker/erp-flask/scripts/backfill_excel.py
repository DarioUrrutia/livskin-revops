"""Backfill simplificado: importa Datos_Livskin.xlsx -> Postgres.

Versión MVP de ADR-0025 sin container Docker dedicado. Preserva los
códigos originales (LIVCLIENT####, LIVTRAT####, LIVPAGO####) del Excel.

Uso (dentro del container erp-flask):
    python /app/scripts/backfill_excel.py /data/datos_livskin.xlsx
"""
import hashlib
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

import openpyxl

sys.path.insert(0, "/app")

from db import session_scope
from models.cliente import Cliente
from models.gasto import Gasto
from models.pago import Pago
from models.venta import Venta


PHONE_RE = re.compile(r"[\s\-\(\)\.]")


def norm_phone(raw: Any) -> Optional[str]:
    if raw is None or str(raw).strip() == "":
        return None
    s = PHONE_RE.sub("", str(raw).strip())
    if not s:
        return None
    if s.startswith("00"):
        s = "+" + s[2:]
    if not s.startswith("+") and s.isdigit():
        if len(s) == 9 and s[0] in "9876":
            s = "+51" + s
    digits = s.lstrip("+")
    if not digits.isdigit() or len(digits) < 7 or len(digits) > 15:
        return None
    return s if s.startswith("+") else "+" + s


def norm_email(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    return s if s and "@" in s else None


def parse_date(raw: Any) -> Optional[date]:
    if raw is None or str(raw).strip() == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def to_dec(raw: Any) -> Optional[Decimal]:
    if raw is None or str(raw).strip() == "":
        return None
    try:
        return Decimal(str(raw).replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def hash_email(email: str) -> str:
    return hashlib.sha256(email.encode("utf-8")).hexdigest()


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    print(f"Excel cargado: {xlsx_path}")
    print(f"Hojas: {wb.sheetnames}")

    inserted = {"clientes": 0, "ventas": 0, "pagos": 0, "gastos": 0}
    errors: list[str] = []

    with session_scope() as db:
        # ============================
        # CLIENTES (hoja "Clientes")
        # ============================
        ws = wb["Clientes"]
        rows = list(ws.iter_rows(values_only=True, min_row=2))
        for i, row in enumerate(rows):
            if not row or not row[0]:
                continue
            try:
                cod = str(row[0]).strip()
                if not cod.startswith("LIVCLIENT"):
                    continue
                nombre = (str(row[1]) if row[1] else "").strip() or "(sin nombre)"
                tel = norm_phone(row[2])
                fecha_nac = parse_date(row[3])
                fecha_reg = parse_date(row[4])
                email = norm_email(row[5])
                cliente = Cliente(
                    cod_cliente=cod,
                    nombre=nombre,
                    phone_e164=tel,
                    email_lower=email,
                    email_hash_sha256=hash_email(email) if email else None,
                    fecha_nacimiento=fecha_nac,
                    fecha_registro=fecha_reg,
                    fuente="organico",
                    canal_adquisicion="legacy",
                    consent_marketing=False,
                    activo=True,
                )
                db.add(cliente)
                inserted["clientes"] += 1
            except Exception as e:
                errors.append(f"Cliente row {i + 2}: {e}")
        db.flush()
        print(f"Clientes insertados: {inserted['clientes']}")

        # ============================
        # VENTAS (hoja "Ventas") — flat 1-fila-1-item
        # ============================
        ws = wb["Ventas"]
        rows = list(ws.iter_rows(values_only=True, min_row=2))
        cod_cliente_set = {c.cod_cliente for c in db.query(Cliente).all()}

        for i, row in enumerate(rows):
            if not row or row[1] is None:
                continue
            try:
                num_seq = int(row[0]) if row[0] is not None else None
                fecha = parse_date(row[1])
                cod_cli = (str(row[2]) if row[2] else "").strip()
                cliente_nom = (str(row[3]) if row[3] else "").strip()
                cliente_tel = (str(row[4]) if row[4] else "").strip() or None
                tipo = (str(row[5]) if row[5] else "").strip() or "Tratamiento"
                cod_item = (str(row[6]) if row[6] else "").strip()
                if not cod_item:
                    continue
                categoria = (str(row[7]) if row[7] else "").strip() or None
                zona = (str(row[8]) if row[8] else "").strip() or None
                proxima_cita = parse_date(row[9])
                fecha_nac = parse_date(row[10])
                moneda = (str(row[11]) if row[11] else "Soles").strip()
                if moneda.lower().startswith("sol"):
                    moneda = "PEN"
                total = to_dec(row[12]) or Decimal("0")
                efectivo = to_dec(row[13])
                yape = to_dec(row[14])
                plin = to_dec(row[15])
                giro = to_dec(row[16])
                debe_excel = to_dec(row[17])
                pagado_excel = to_dec(row[18])
                tc = to_dec(row[19])
                precio_lista = to_dec(row[20])
                descuento = to_dec(row[21]) or Decimal("0")

                if cod_cli not in cod_cliente_set:
                    errors.append(f"Venta row {i + 2}: cod_cliente {cod_cli} no existe")
                    continue

                if fecha is None:
                    errors.append(f"Venta row {i + 2}: fecha inválida")
                    continue

                venta = Venta(
                    num_secuencial=num_seq,
                    fecha=fecha,
                    cod_cliente=cod_cli,
                    cliente_nombre=cliente_nom,
                    cliente_telefono=cliente_tel,
                    tipo=tipo if tipo in ("Tratamiento", "Producto", "Certificado", "Promocion") else "Tratamiento",
                    cod_item=cod_item,
                    categoria=categoria,
                    zona_cantidad_envase=zona,
                    proxima_cita=proxima_cita,
                    fecha_nac_cliente=fecha_nac,
                    moneda=moneda,
                    total=total,
                    efectivo=efectivo,
                    yape=yape,
                    plin=plin,
                    giro=giro,
                    debe=debe_excel,
                    pagado=pagado_excel,
                    tc=tc,
                    precio_lista=precio_lista,
                    descuento=descuento,
                )
                db.add(venta)
                inserted["ventas"] += 1
            except Exception as e:
                errors.append(f"Venta row {i + 2}: {e}")
        db.flush()
        print(f"Ventas insertadas: {inserted['ventas']}")

        # ============================
        # PAGOS (hoja "Pagos")
        # ============================
        ws = wb["Pagos"]
        rows = list(ws.iter_rows(values_only=True, min_row=2))
        for i, row in enumerate(rows):
            if not row or row[1] is None:
                continue
            try:
                num_seq = int(row[0]) if row[0] is not None else None
                fecha = parse_date(row[1])
                cod_cli = (str(row[2]) if row[2] else "").strip()
                cliente_nom = (str(row[3]) if row[3] else "").strip()
                monto = to_dec(row[4]) or Decimal("0")
                efectivo = to_dec(row[5])
                yape = to_dec(row[6])
                plin = to_dec(row[7])
                giro = to_dec(row[8])
                notas = (str(row[9]) if row[9] else "").strip() or None
                cod_item = (str(row[10]) if row[10] else "").strip() or None
                categoria = (str(row[11]) if row[11] else "").strip() or None
                cod_pago = (str(row[12]) if row[12] else "").strip()
                if not cod_pago:
                    continue
                if cod_cli not in cod_cliente_set:
                    errors.append(f"Pago row {i + 2}: cod_cliente {cod_cli} no existe")
                    continue
                if fecha is None:
                    errors.append(f"Pago row {i + 2}: fecha inválida")
                    continue

                # Inferir tipo_pago desde notas/categoria del Excel
                tipo_pago = "normal"
                if notas:
                    n_lower = notas.lower()
                    if "crédito aplicado" in n_lower or "credito aplicado" in n_lower:
                        tipo_pago = "credito_aplicado"
                    elif "abono" in n_lower:
                        tipo_pago = "abono_deuda"
                if categoria and (
                    categoria.upper().startswith("CRÉDITO") or categoria.upper().startswith("CREDITO")
                    or categoria.upper().startswith("ANTICIPO")
                ):
                    tipo_pago = "credito_generado"

                pago = Pago(
                    num_secuencial=num_seq,
                    cod_pago=cod_pago,
                    fecha=fecha,
                    cod_cliente=cod_cli,
                    cliente_nombre=cliente_nom,
                    cod_item=cod_item,
                    categoria=categoria,
                    monto=monto,
                    efectivo=efectivo,
                    yape=yape,
                    plin=plin,
                    giro=giro,
                    tipo_pago=tipo_pago,
                    notas=notas,
                )
                db.add(pago)
                inserted["pagos"] += 1
            except Exception as e:
                errors.append(f"Pago row {i + 2}: {e}")
        db.flush()
        print(f"Pagos insertados: {inserted['pagos']}")

        # ============================
        # GASTOS (hoja "Gastos") — vacía típicamente
        # ============================
        ws = wb["Gastos"]
        rows = list(ws.iter_rows(values_only=True, min_row=2))
        for i, row in enumerate(rows):
            if not row or not any(c for c in row):
                continue
            try:
                num_seq = int(row[0]) if row[0] is not None else None
                fecha = parse_date(row[1])
                tipo = (str(row[2]) if row[2] else "").strip() or None
                desc = (str(row[3]) if row[3] else "").strip() or None
                dest = (str(row[4]) if row[4] else "").strip() or None
                monto = to_dec(row[5]) or Decimal("0")
                metodo = (str(row[6]) if row[6] else "").strip() or None
                if fecha is None:
                    continue
                gasto = Gasto(
                    num_secuencial=num_seq,
                    fecha=fecha,
                    tipo=tipo,
                    descripcion=desc,
                    destinatario=dest,
                    monto=monto,
                    metodo_pago=metodo,
                )
                db.add(gasto)
                inserted["gastos"] += 1
            except Exception as e:
                errors.append(f"Gasto row {i + 2}: {e}")
        db.flush()
        print(f"Gastos insertados: {inserted['gastos']}")

    print()
    print("=== RESUMEN ===")
    for k, v in inserted.items():
        print(f"  {k}: {v}")
    print(f"  errores: {len(errors)}")
    if errors:
        print()
        print("Primeros 10 errores:")
        for e in errors[:10]:
            print(f"  - {e}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python backfill_excel.py /path/to/excel.xlsx", file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1])
